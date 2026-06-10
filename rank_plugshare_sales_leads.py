"""
Build a sales-friendly ranked workbook from PlugShare unavailable/maintenance rows.

The WAIRE lead ranker used a simple hybrid idea:
  - pain signal: how much visible urgency does the account have?
  - credibility signal: does Integra already win in this kind of organization?

For PlugShare, pain is broken/under-repair charger severity, affected port count,
and charging speed. Credibility is inferred from the same customer-CSV industry
buckets captured in the WAIRE project notes.
"""
from __future__ import annotations

import argparse
import csv
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


TARGET_STATUSES = {"OUTOFORDER", "UNDER_REPAIR", "UNAVAILABLE"}

US_ZIP_RANGES = {
    "CA": (90000, 96199),
    "CT": (6000, 6999),
    "ME": (3900, 4999),
    "MA": (1000, 2799),
    "NH": (3000, 3899),
    "NJ": (7000, 8999),
    "NY": (10000, 14999),
    "PA": (15000, 19699),
    "RI": (2800, 2999),
    "VT": (5000, 5999),
}

STATUS_LABELS = {
    "OUTOFORDER": "Out of order",
    "UNDER_REPAIR": "Under repair",
    "UNAVAILABLE": "Unavailable",
}

STATUS_WEIGHTS = {
    "OUTOFORDER": 3.0,
    "UNDER_REPAIR": 4.0,
    "UNAVAILABLE": 3.0,
}

# Grounded in memory/projects/ai_lead_gen/customer_icp_analysis.md.
INDUSTRY_CUSTOMER_COUNTS = {
    "Hotel / Hospitality": 26,
    "Real Estate / Property Mgmt": 21,
    "Multifamily / Apartments": 15,
    "Auto Dealership / Automotive": 14,
    "Manufacturing / Industrial": 14,
    "Education": 10,
    "Municipal / Government": 9,
    "Healthcare": 9,
    "Retail / Commercial": 6,
    "Fuel / Gas Station": 4,
    "Food / Beverage Distribution": 4,
    "Logistics / 3PL / Trucking": 0,
    "Packaging / Containers": 0,
    "(uncategorized)": 0,
}

# First match wins. Keep more-specific categories above fuzzier categories.
BUCKETS = {
    "Auto Dealership / Automotive": [
        "chrysler", "dodge", "jeep", "ram", "chevrolet", "chevy", "honda",
        "nissan", "toyota", "ford", "mazda", "kia", "hyundai", "subaru",
        "volkswagen", "audi", "bmw", "mercedes", "lexus", "cadillac",
        "buick", "gmc", "dealer", "dealership", "auto group", "motors",
        "cars", "car ", "ferrario", "lia ", "della", "denooyer",
    ],
    "Hotel / Hospitality": [
        "hotel", "motel", " inn", "inn ", "resort", "lodge", "hospitality",
        "marriott", "hilton", "hampton", "courtyard", "hyatt", "sheraton",
        "holiday inn", "best western", "comfort", "sleep inn", "suites",
        "restaurant", "tavern", "brew", "cafe", "diner", "casino",
    ],
    "Multifamily / Apartments": [
        "apartments", "apartment", "condo", "condominium", "villas",
        "residences", "residence", "senior living", "senior housing",
        "manor", "lofts", "townhomes", "homes",
    ],
    "Real Estate / Property Mgmt": [
        "realty", "property", "properties", "real estate", "holdings",
        "equities", "management", "development", "plaza", "mall",
        "parking", "garage", "office", "center", "centre", "llc",
    ],
    "Education": [
        "suny", "university", "college", "school", "academy", "campus",
        "institute", "education",
    ],
    "Municipal / Government": [
        "town of", "city of", "village of", "county", "public works",
        "authority", "state park", "park -", "state office", "municipal",
        "library", "courthouse", "transit", "airport", "mta", "nys",
    ],
    "Healthcare": [
        "hospital", "medical", "health", "clinic", "urgent care",
        "dermatology", "dental", "veterinary", "rehab",
    ],
    "Fuel / Gas Station": [
        "fuel", "gas station", "petroleum", "travel plaza", "service area",
        "truck stop", "pilot", "shell", "mobil", "exxon", "sunoco",
    ],
    "Food / Beverage Distribution": [
        "food", "foods", "beverage", "restaurant depot", "market",
        "supermarket", "grocery", "brewery",
    ],
    "Retail / Commercial": [
        "target", "walmart", "wal-mart", "home depot", "lowe", "costco",
        "shop", "shopping", "store", "marketplace", "retail", "bank",
        "credit union", "museum", "cinema", "theater",
    ],
    "Manufacturing / Industrial": [
        "manufacturing", "industrial", "factory", "plant", "works",
        "power", "energy", "solar",
    ],
    "Logistics / 3PL / Trucking": [
        "logistics", "trucking", "freight", "transport", "warehouse",
        "distribution", "shipping",
    ],
    "Packaging / Containers": [
        "packaging", "container",
    ],
}


@dataclass
class RankedLocation:
    row: dict[str, object]
    raw_rows: list[dict[str, str]]


def clean_float(value: str | None) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except ValueError:
        return 0.0


def norm_status(value: str | None) -> str:
    return (value or "").strip().upper()


def categorize(name: str, address: str) -> tuple[str, str]:
    haystack = f"{name or ''} {address or ''}".lower()
    for category, keywords in BUCKETS.items():
        hits = [kw for kw in keywords if kw in haystack]
        if hits:
            return category, ", ".join(hits[:3])
    return "(uncategorized)", ""


def extract_last_zip(address: str) -> str:
    """Return the last 5-digit ZIP-like token.

    We use the last candidate instead of the first because many addresses start
    with a 5-digit street number or highway number.
    """
    if not address:
        return ""
    matches = re.findall(r"(?<!\d)(\d{5})(?!\d)", address)
    return matches[-1] if matches else ""


def parse_city_zip(address: str) -> tuple[str, str]:
    if not address:
        return "", ""
    zip_code = extract_last_zip(address)
    parts = [p.strip() for p in address.split(",")]
    city = ""
    if len(parts) >= 4:
        city = parts[-3]
    elif len(parts) == 3:
        city = parts[-2]
    elif len(parts) >= 2:
        tail = re.sub(r"\b[A-Z]{2}\b", "", parts[-1])
        tail = re.sub(r"\b\d{5}(?:-\d{4})?\b", "", tail).strip()
        city = tail
    city = re.sub(r"\b[A-Z]{2}\b", "", city)
    city = re.sub(r"\b\d{5}(?:-\d{4})?\b", "", city)
    city = re.sub(r"\s+", " ", city).strip(" ,")
    return city, zip_code


def state_matches(row: dict[str, str], state_filter: str | None) -> bool:
    if not state_filter:
        return True
    state = state_filter.upper()
    address = row.get("address", "")
    zip_code = extract_last_zip(address)
    if zip_code and state in US_ZIP_RANGES:
        lo, hi = US_ZIP_RANGES[state]
        return lo <= int(zip_code) <= hi
    address_has_state = bool(re.search(rf"\b{re.escape(state)}\b", address, re.IGNORECASE))
    if state == "NY":
        address_has_state = address_has_state or bool(re.search(r"\bNew York\b", address, re.IGNORECASE))
    return (row.get("state_or_province") or "").upper() == state or address_has_state


def summarize_status(counter: Counter[str]) -> str:
    pieces = []
    for status in ["OUTOFORDER", "UNDER_REPAIR", "UNAVAILABLE"]:
        if counter.get(status, 0):
            pieces.append(f"{counter[status]} {STATUS_LABELS[status]}")
    return ", ".join(pieces)


def network_tier(network_name: str) -> str:
    if not network_name:
        return "Verify network before pitch"
    if "chargepoint" in network_name.lower():
        return "ChargePoint - rip-and-replace"
    return "Non-ChargePoint - takeover candidate"


def score_location(rows: list[dict[str, str]]) -> RankedLocation:
    sample = rows[0]
    statuses = Counter(norm_status(r.get("status")) for r in rows)
    target_rows = [r for r in rows if norm_status(r.get("status")) in TARGET_STATUSES]
    affected_ports = len({r.get("outlet_id") or f"row-{i}" for i, r in enumerate(target_rows)})
    affected_stations = len({r.get("station_id") for r in target_rows if r.get("station_id")})
    max_kw = max([clean_float(r.get("kilowatts")) for r in target_rows] or [0.0])
    connectors = sorted({r.get("connector", "").strip() for r in target_rows if r.get("connector", "").strip()})
    network_names = [
        r.get("network_name") or r.get("majority_network_name") or ""
        for r in target_rows
    ]
    network_name_counts = Counter(name for name in network_names if name)
    primary_network = network_name_counts.most_common(1)[0][0] if network_name_counts else ""
    takeover_tier = network_tier(primary_network)

    category, keyword_hits = categorize(sample.get("name", ""), sample.get("address", ""))
    customer_count = INDUSTRY_CUSTOMER_COUNTS.get(category, 0)

    weighted_issues = sum(statuses[s] * STATUS_WEIGHTS.get(s, 0.0) for s in TARGET_STATUSES)
    pain_score = min(45.0, 15.0 * math.log1p(weighted_issues))
    speed_score = 0.0
    if max_kw >= 150:
        speed_score = 15.0
    elif max_kw >= 100:
        speed_score = 12.0
    elif max_kw >= 50:
        speed_score = 9.0
    elif max_kw >= 19:
        speed_score = 5.0
    port_score = min(10.0, affected_ports * 1.5 + affected_stations * 1.0)
    credibility_score = min(25.0, 7.5 * math.log1p(customer_count))
    phone_score = 10.0 if sample.get("phone_number") else 0.0

    total_score = min(100.0, pain_score + speed_score + port_score + credibility_score + phone_score)
    if total_score >= 75:
        tier = "A - Call first"
    elif total_score >= 55:
        tier = "B - Strong follow-up"
    else:
        tier = "C - Lower priority"

    reasons = []
    if sample.get("phone_number"):
        reasons.append("phone found")
    if summarize_status(statuses):
        reasons.append(summarize_status(statuses))
    if max_kw >= 50:
        reasons.append(f"fast charging up to {max_kw:g} kW")
    if customer_count:
        reasons.append(f"{category} is an Integra-proven segment")
    elif category != "(uncategorized)":
        reasons.append(f"{category} match, but limited existing-customer proof")
    else:
        reasons.append("needs manual industry review")
    if takeover_tier.startswith("Non-ChargePoint"):
        reasons.append("network is a software takeover candidate")
    elif takeover_tier.startswith("ChargePoint"):
        reasons.append("ChargePoint means rip-and-replace pitch")
    else:
        reasons.append("verify network before takeover pitch")

    city, zip_code = parse_city_zip(sample.get("address", ""))
    row = {
        "Priority Rank": 0,
        "Score": round(total_score, 1),
        "Tier": tier,
        "Call First Reason": "; ".join(reasons),
        "Location / Business": sample.get("name", ""),
        "Phone": sample.get("phone_number", ""),
        "Address": sample.get("address", ""),
        "City": city,
        "ZIP": zip_code,
        "State": sample.get("state_or_province", ""),
        "Status Summary": summarize_status(statuses),
        "Affected Ports": affected_ports,
        "Affected Stations": affected_stations,
        "Max kW": max_kw,
        "Connector Types": ", ".join(connectors),
        "Network": primary_network,
        "OCPP Pitch Tier": takeover_tier,
        "ICP Segment": category,
        "ICP Customer Count": customer_count,
        "Matched Keywords": keyword_hits,
        "Phone Source": sample.get("phone_source", ""),
        "Google Maps": sample.get("google_maps_url", ""),
        "PlugShare": sample.get("url", ""),
        "Location ID": sample.get("location_id", ""),
        "Pain Score": round(pain_score, 1),
        "Speed Score": round(speed_score, 1),
        "Port Score": round(port_score, 1),
        "ICP Score": round(credibility_score, 1),
        "Phone Score": round(phone_score, 1),
    }
    return RankedLocation(row=row, raw_rows=rows)


def load_ranked(input_csv: Path, state_filter: str | None = None) -> list[RankedLocation]:
    with input_csv.open(newline="", encoding="utf-8-sig") as f:
        rows = [r for r in csv.DictReader(f) if norm_status(r.get("status")) in TARGET_STATUSES]

    groups: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        key = row.get("location_id") or f"{row.get('name')}|{row.get('address')}"
        groups[key].append(row)

    if state_filter:
        groups = {
            key: group_rows
            for key, group_rows in groups.items()
            if state_matches(group_rows[0], state_filter)
        }

    ranked = [score_location(group_rows) for group_rows in groups.values()]
    ranked.sort(
        key=lambda item: (
            item.row["Score"],
            1 if item.row["Phone"] else 0,
            item.row["Affected Ports"],
            item.row["Max kW"],
            item.row["Location / Business"],
        ),
        reverse=True,
    )
    for i, item in enumerate(ranked, 1):
        item.row["Priority Rank"] = i
    return ranked


def append_table(ws, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def apply_table_style(ws, freeze: str = "A2") -> None:
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="E7E6E6"))


def set_widths(ws, widths: dict[str, int], default: int = 14) -> None:
    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = widths.get(letter, default)


def build_workbook(ranked: list[RankedLocation], output_xlsx: Path) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    call_headers = list(ranked[0].row.keys()) if ranked else []
    call_rows = [r.row for r in ranked]
    status_counts = Counter()
    raw_count = 0
    phone_locations = 0
    tiers = Counter()
    for item in ranked:
        tiers[item.row["Tier"]] += 1
        if item.row["Phone"]:
            phone_locations += 1
        for raw in item.raw_rows:
            raw_count += 1
            status_counts[norm_status(raw.get("status"))] += 1

    summary_rows = [
        ["Metric", "Value"],
        ["Ranked locations", len(ranked)],
        ["Raw charger rows", raw_count],
        ["Locations with phone", phone_locations],
        ["A - Call first", tiers.get("A - Call first", 0)],
        ["B - Strong follow-up", tiers.get("B - Strong follow-up", 0)],
        ["C - Lower priority", tiers.get("C - Lower priority", 0)],
        ["Out of order rows", status_counts.get("OUTOFORDER", 0)],
        ["Under repair rows", status_counts.get("UNDER_REPAIR", 0)],
        ["Unavailable rows", status_counts.get("UNAVAILABLE", 0)],
    ]
    for row in summary_rows:
        summary.append(row)
    summary["D1"] = "How to use this"
    summary["D2"] = (
        "Start with the Call List tab. It is sorted by Score so the best-fit, "
        "highest-pain, easiest-to-call locations are first."
    )
    summary["D4"] = "Score recipe"
    summary["D5"] = "Broken/repair severity + charger speed + affected port count + ICP fit + phone availability."
    summary["D7"] = "ICP source"
    summary["D8"] = "Integra customer analysis from the WAIRE lead-gen project, adapted to PlugShare locations."
    apply_table_style(summary)
    set_widths(summary, {"A": 24, "B": 16, "D": 48})
    summary["D2"].alignment = Alignment(wrap_text=True, vertical="top")
    summary["D5"].alignment = Alignment(wrap_text=True, vertical="top")
    summary["D8"].alignment = Alignment(wrap_text=True, vertical="top")

    call_ws = wb.create_sheet("Call List")
    append_table(call_ws, call_headers, call_rows)
    apply_table_style(call_ws)
    set_widths(call_ws, {
        "A": 12, "B": 10, "C": 20, "D": 48, "E": 34, "F": 18,
        "G": 44, "H": 18, "I": 10, "J": 9, "K": 28, "L": 14,
        "M": 16, "N": 10, "O": 18, "P": 28, "Q": 16, "R": 20,
        "S": 15, "T": 18, "U": 18, "V": 14, "W": 11, "X": 11,
        "Y": 11, "Z": 10, "AA": 12,
    })
    for row_idx in range(2, call_ws.max_row + 1):
        tier = call_ws.cell(row_idx, 3).value
        fill = None
        if tier == "A - Call first":
            fill = PatternFill("solid", fgColor="C6EFCE")
        elif tier == "B - Strong follow-up":
            fill = PatternFill("solid", fgColor="FFEB9C")
        elif tier == "C - Lower priority":
            fill = PatternFill("solid", fgColor="FCE4D6")
        if fill:
            for col_idx in range(1, call_ws.max_column + 1):
                call_ws.cell(row_idx, col_idx).fill = fill
        maps_cell = call_ws.cell(row_idx, 20)
        if maps_cell.value:
            maps_cell.hyperlink = maps_cell.value
            maps_cell.style = "Hyperlink"
            maps_cell.value = "Open Maps"
        plug_cell = call_ws.cell(row_idx, 21)
        if plug_cell.value:
            plug_cell.hyperlink = plug_cell.value
            plug_cell.style = "Hyperlink"
            plug_cell.value = "Open PlugShare"

    raw_ws = wb.create_sheet("Raw Charger Rows")
    raw_headers = list(ranked[0].raw_rows[0].keys()) if ranked and ranked[0].raw_rows else []
    raw_rows = []
    for item in ranked:
        for raw in item.raw_rows:
            with_rank = dict(raw)
            with_rank["priority_rank"] = item.row["Priority Rank"]
            with_rank["score"] = item.row["Score"]
            with_rank["tier"] = item.row["Tier"]
            raw_rows.append(with_rank)
    raw_headers = ["priority_rank", "score", "tier"] + raw_headers
    append_table(raw_ws, raw_headers, raw_rows)
    apply_table_style(raw_ws)
    set_widths(raw_ws, {"A": 12, "B": 10, "C": 20, "D": 14, "E": 34, "F": 44, "G": 18, "H": 18}, default=16)

    notes = wb.create_sheet("Scoring Notes")
    notes_rows = [
        ["Component", "Max Points", "Why it matters"],
        ["Pain Score", 45, "More broken/under-repair ports means a more urgent operational issue."],
        ["Speed Score", 15, "DC fast chargers and higher-kW locations are more valuable and more painful when down."],
        ["Port Score", 10, "Multiple affected ports/stations indicate a bigger site-level problem."],
        ["ICP Score", 25, "Boosts segments where Integra already has customer proof."],
        ["Phone Score", 10, "A found phone number makes the lead immediately callable."],
    ]
    for row in notes_rows:
        notes.append(row)
    notes.append([])
    notes.append(["ICP Segment", "Customer Count Used", "Source"])
    for segment, count in sorted(INDUSTRY_CUSTOMER_COUNTS.items(), key=lambda x: (-x[1], x[0])):
        notes.append([segment, count, "Integra customer ICP analysis / WAIRE ranker"])
    apply_table_style(notes)
    set_widths(notes, {"A": 32, "B": 18, "C": 72})

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def write_csv(ranked: list[RankedLocation], output_csv: Path) -> None:
    if not ranked:
        return
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    headers = list(ranked[0].row.keys())
    with output_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows([r.row for r in ranked])


def main() -> int:
    parser = argparse.ArgumentParser(description="Rank PlugShare unavailable/maintenance locations for sales outreach.")
    parser.add_argument("--input", required=True, help="Input CSV from filter/enrichment pipeline.")
    parser.add_argument("--output-xlsx", required=True, help="Polished ranked workbook to write.")
    parser.add_argument("--output-csv", help="Optional ranked call-list CSV to write.")
    parser.add_argument("--state-filter", help="Optional state/province filter, e.g. NY. For NY, ZIP is used when present.")
    args = parser.parse_args()

    ranked = load_ranked(Path(args.input), args.state_filter)
    if not ranked:
        raise SystemExit("No target status rows found.")
    build_workbook(ranked, Path(args.output_xlsx))
    if args.output_csv:
        write_csv(ranked, Path(args.output_csv))

    status_counts = Counter()
    for item in ranked:
        for raw in item.raw_rows:
            status_counts[norm_status(raw.get("status"))] += 1
    print(f"Ranked locations: {len(ranked)}")
    print(f"Raw target charger rows: {sum(status_counts.values())}")
    print(f"Rows by status: {dict(status_counts)}")
    print(f"Wrote: {args.output_xlsx}")
    if args.output_csv:
        print(f"Wrote: {args.output_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
