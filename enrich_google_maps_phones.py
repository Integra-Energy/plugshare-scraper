#!/usr/bin/env python3
"""Enrich NY unavailable charger rows with Google Maps phone numbers.

This drives a visible Google Maps browser session and searches by PlugShare
business name + address. Results are cached by location_id so interrupted runs can
resume without re-querying completed locations.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import time
import urllib.parse
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError, sync_playwright

TARGET_STATUSES = {"OUTOFORDER", "UNDER_REPAIR", "UNAVAILABLE"}
PHONE_RE = re.compile(r"(?:\+?1[\s.-]?)?\(?\d{3}\)?[\s.-]\d{3}[\s.-]\d{4}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Use Google Maps to enrich NY charger phone numbers.")
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("north-america-all-chargers-by-state-province.csv"),
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=Path("ny-out-of-order-unavailable-chargers-google-maps-phones.xlsx"),
    )
    parser.add_argument(
        "--output-csv",
        type=Path,
        default=Path("ny-out-of-order-unavailable-chargers-google-maps-phones.csv"),
    )
    parser.add_argument(
        "--cache",
        type=Path,
        default=Path("ny-google-maps-phone-cache.json"),
    )
    parser.add_argument("--delay", type=float, default=2.0)
    parser.add_argument("--wait-ms", type=int, default=4500, help="Milliseconds to wait after each Maps search.")
    parser.add_argument("--limit", type=int, help="Optional number of unique locations to query.")
    return parser.parse_args()


def load_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    rows: list[dict[str, str]] = []
    with path.open(encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        headers = reader.fieldnames or []
        for row in reader:
            if row.get("state_or_province") == "NY" and row.get("status") in TARGET_STATUSES:
                rows.append(row)
    return headers, rows


def unique_locations(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[str] = set()
    locations: list[dict[str, str]] = []
    for row in rows:
        location_id = row["location_id"]
        if location_id in seen:
            continue
        seen.add(location_id)
        locations.append(row)
    return locations


def load_cache(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_cache(path: Path, cache: dict[str, dict[str, str]]) -> None:
    path.write_text(json.dumps(cache, indent=2, sort_keys=True), encoding="utf-8")


def extract_phone(page) -> str:
    selectors = [
        "button[data-item-id^='phone:tel:']",
        "[data-item-id^='phone:tel:']",
        "[aria-label^='Phone:']",
    ]
    for selector in selectors:
        locator = page.locator(selector)
        count = locator.count()
        for index in range(count):
            element = locator.nth(index)
            label = element.get_attribute("aria-label") or element.inner_text(timeout=1500)
            match = PHONE_RE.search(label or "")
            if match:
                return match.group(0)
    try:
        body_text = page.locator("body").inner_text(timeout=5000)
    except PlaywrightTimeoutError:
        return ""
    matches = PHONE_RE.findall(body_text)
    return matches[0] if matches else ""


def normalize_e164(phone: str) -> str:
    digits = re.sub(r"\D", "", phone)
    if len(digits) == 10:
        return "+1" + digits
    if len(digits) == 11 and digits.startswith("1"):
        return "+" + digits
    return ""


def query_google_maps(
    locations: list[dict[str, str]],
    cache: dict[str, dict[str, str]],
    delay: float,
    wait_ms: int,
    limit: int | None,
) -> None:
    pending = [location for location in locations if location["location_id"] not in cache]
    if limit is not None:
        pending = pending[:limit]
    print(f"Google Maps locations to query: {len(pending)}")
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False)
        page = browser.new_page(viewport={"width": 1400, "height": 900})
        for index, location in enumerate(pending, start=1):
            location_id = location["location_id"]
            query = f"{location['name']} {location['address']} phone"
            url = "https://www.google.com/maps/search/" + urllib.parse.quote(query)
            result = {
                "phone_number": "",
                "phone_e164": "",
                "phone_source": "",
                "google_maps_url": "",
                "query": query,
            }
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=30_000)
                page.wait_for_timeout(wait_ms)
                phone = extract_phone(page)
                if phone:
                    result["phone_number"] = phone
                    result["phone_e164"] = normalize_e164(phone)
                    result["phone_source"] = "Google Maps"
                result["google_maps_url"] = page.url
            except Exception as exc:
                result["error"] = str(exc)
            cache[location_id] = result
            if index % 10 == 0 or index == len(pending):
                print(f"Queried {index}/{len(pending)}; phones found: {sum(1 for value in cache.values() if value.get('phone_number'))}")
                save_cache(Path("ny-google-maps-phone-cache.json"), cache)
            time.sleep(delay)
        browser.close()


def write_outputs(headers: list[str], rows: list[dict[str, str]], cache: dict[str, dict[str, str]], output_csv: Path, output_xlsx: Path) -> None:
    output_headers = headers[:]
    insert_after = output_headers.index("address") + 1
    for column in ["phone_number", "phone_e164", "phone_source", "google_maps_url"]:
        if column not in output_headers:
            output_headers.insert(insert_after, column)
            insert_after += 1
    for row in rows:
        details = cache.get(row["location_id"], {})
        for column in ["phone_number", "phone_e164", "phone_source", "google_maps_url"]:
            row[column] = details.get(column, "")
    rows.sort(
        key=lambda row: (
            not bool(row.get("phone_number")),
            row.get("address", ""),
            row.get("name", ""),
            row.get("location_id", ""),
            row.get("station_id", ""),
            row.get("outlet_id", ""),
        )
    )
    with output_csv.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=output_headers)
        writer.writeheader()
        writer.writerows(rows)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    data = workbook.create_sheet("NY Chargers")
    fill = PatternFill("solid", fgColor="174A7E")
    font = Font(bold=True, color="FFFFFF")
    status_counts = Counter(row["status"] for row in rows)
    locations_with_phone = len({row["location_id"] for row in rows if row.get("phone_number")})
    rows_with_phone = sum(1 for row in rows if row.get("phone_number"))
    summary_rows = [
        ["Metric", "Value"],
        ["State", "NY"],
        ["Included statuses", ", ".join(sorted(TARGET_STATUSES))],
        ["Total charger outlet rows", len(rows)],
        ["Unique affected locations", len({row["location_id"] for row in rows})],
        ["Rows with phone number", rows_with_phone],
        ["Locations with phone number", locations_with_phone],
        ["OUTOFORDER", status_counts.get("OUTOFORDER", 0)],
        ["UNDER_REPAIR", status_counts.get("UNDER_REPAIR", 0)],
        ["UNAVAILABLE", status_counts.get("UNAVAILABLE", 0)],
    ]
    for row in summary_rows:
        summary.append(row)
    for cell in summary[1]:
        cell.fill = fill
        cell.font = font
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = f"A1:B{len(summary_rows)}"
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 60

    data.append(output_headers)
    for cell in data[1]:
        cell.fill = fill
        cell.font = font
    for row in rows:
        data.append([row.get(header, "") for header in output_headers])
    data.freeze_panes = "A2"
    data.auto_filter.ref = f"A1:{get_column_letter(len(output_headers))}{len(rows) + 1}"
    widths = [14, 30, 54, 18, 18, 24, 48, 15, 12, 12, 20, 14, 14, 12, 16, 12, 42]
    for index, width in enumerate(widths[: len(output_headers)], start=1):
        data.column_dimensions[get_column_letter(index)].width = width
    workbook.save(output_xlsx)
    check = load_workbook(output_xlsx, read_only=True)
    verified_rows = sum(1 for _ in check["NY Chargers"].iter_rows(values_only=True)) - 1
    check.close()
    print(f"Wrote {output_xlsx.resolve()}")
    print(f"Wrote {output_csv.resolve()}")
    print(f"Verified workbook rows: {verified_rows}")
    print(f"Rows with phone: {rows_with_phone}/{len(rows)}")
    print(f"Locations with phone: {locations_with_phone}/{len({row['location_id'] for row in rows})}")


def main() -> int:
    args = parse_args()
    headers, rows = load_rows(args.input)
    cache = load_cache(args.cache)
    locations = unique_locations(rows)
    query_google_maps(locations, cache, args.delay, args.wait_ms, args.limit)
    save_cache(args.cache, cache)
    write_outputs(headers, rows, cache, args.output_csv, args.output_xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
