#!/usr/bin/env python3
"""Build a memory-efficient Excel workbook from the sorted PlugShare CSV."""

from __future__ import annotations

import argparse
import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="174A7E")
HEADER_FONT = Font(bold=True, color="FFFFFF")
COLUMN_WIDTHS = (14, 30, 54, 15, 12, 12, 20, 14, 14, 12, 16, 12, 42)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a state/province-tabbed PlugShare workbook.")
    parser.add_argument("input_csv", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    return parser.parse_args()


def subdivision(row: dict[str, str]) -> str:
    return row.get("state_or_province") or "Unknown"


def styled_header(worksheet, headers: list[str]) -> None:
    cells = []
    for value in headers:
        cell = WriteOnlyCell(worksheet, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cells.append(cell)
    worksheet.append(cells)
    worksheet.freeze_panes = "A2"
    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width


def add_summary(workbook: Workbook, counts: Counter[str]) -> None:
    sheet = workbook.create_sheet("Summary")
    styled_header(sheet, ["State / Province", "Charger Rows"])
    for name, count in sorted(counts.items(), key=lambda item: (item[0] == "Unknown", item[0])):
        sheet.append([name, count])
    sheet.auto_filter.ref = f"A1:B{len(counts) + 1}"


def main() -> int:
    args = parse_args()
    with args.input_csv.open(encoding="utf-8-sig", newline="") as source:
        counts = Counter(subdivision(row) for row in csv.DictReader(source))

    workbook = Workbook(write_only=True)
    add_summary(workbook, counts)

    current_name = ""
    current_sheet = None
    current_rows = 0
    with args.input_csv.open(encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        headers = reader.fieldnames or []
        for row in reader:
            name = subdivision(row)
            if name != current_name:
                if current_sheet is not None:
                    current_sheet.auto_filter.ref = f"A1:{chr(64 + len(headers))}{current_rows + 1}"
                current_name = name
                current_sheet = workbook.create_sheet(name)
                styled_header(current_sheet, headers)
                current_rows = 0
            current_sheet.append([row.get(header, "") for header in headers])
            current_rows += 1
        if current_sheet is not None:
            current_sheet.auto_filter.ref = f"A1:{chr(64 + len(headers))}{current_rows + 1}"

    args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output_xlsx)
    print(f"Wrote {len(counts)} subdivision sheets plus Summary to {args.output_xlsx.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
